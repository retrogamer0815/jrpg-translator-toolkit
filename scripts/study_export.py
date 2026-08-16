#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Export a Study Library to a formatted, read-only Excel workbook."""

from __future__ import annotations

import argparse
import re
import sqlite3
import sys
from collections import OrderedDict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# The bundled embeddable Python uses an explicit ``._pth`` file, so it does
# not automatically add a launched script's directory to ``sys.path``.
# Keep this bridge self-contained by making its sibling helpers importable.
SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from anki_candidates import (
    candidate_hash,
    decode_optional_hex,
    normalize_japanese,
    parse_vocabulary,
    recommendation_language_context,
    recommendation_profile_signature,
    recommendation_reason_language,
)


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL = PatternFill("solid", fgColor="EAF2F8")
RECOMMENDED_FILL = PatternFill("solid", fgColor="E2F0D9")
MAX_COLUMN_WIDTH = 65
ILLEGAL_XML_CHARACTERS = re.compile(
    "[\x00-\x08\x0B\x0C\x0E-\x1F]"
)
EXCEL_TEXT_LIMIT = 32767
EXCEL_DATETIME_FORMAT = "yyyy-mm-dd hh:mm:ss"
DATETIME_HEADERS = {
    "Date generated",
    "Added to Anki at",
    "AI generated at",
    "Latest",
}


def connect_read_only(path: Path) -> sqlite3.Connection:
    connection = sqlite3.connect(
        f"file:{path.resolve().as_posix()}?mode=ro", uri=True, timeout=10
    )
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA query_only = ON")
    connection.execute("PRAGMA busy_timeout = 10000")
    return connection


def table_exists(connection: sqlite3.Connection, name: str) -> bool:
    return connection.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (name,)
    ).fetchone() is not None


def table_columns(connection: sqlite3.Connection, name: str) -> set[str]:
    if not table_exists(connection, name):
        return set()
    return {
        str(row[1])
        for row in connection.execute(f'PRAGMA table_info("{name}")')
    }


def optional_column(columns: set[str], name: str, fallback: str = "''") -> str:
    return f'e."{name}"' if name in columns else fallback


def selected_explanations(connection: sqlite3.Connection) -> list[dict[str, Any]]:
    explanation_columns = table_columns(connection, "explanations")
    preferred_order = (
        "COALESCE(e.preferred, 0) DESC, e.version DESC"
        if "preferred" in explanation_columns
        else "e.version DESC"
    )
    selected_fields = {
        "key_grammar": optional_column(explanation_columns, "key_grammar"),
        "manual_original_text": optional_column(
            explanation_columns, "manual_original_text"
        ),
        "manually_edited_at": optional_column(
            explanation_columns, "manually_edited_at"
        ),
    }
    rows = connection.execute(
        f"""
        WITH selected AS (
            SELECT e.*,
                   ROW_NUMBER() OVER (
                       PARTITION BY e.group_id ORDER BY {preferred_order}
                   ) AS selection_order
            FROM explanations e
        ), counts AS (
            SELECT group_id, COUNT(*) AS version_count
            FROM explanations GROUP BY group_id
        )
        SELECT g.id AS group_id,
               COALESCE(g.game_profile, '') AS game_profile,
               COALESCE(g.source_japanese, '') AS source_japanese,
               COALESCE(g.created_at, '') AS group_created_at,
               COALESCE(g.updated_at, '') AS group_updated_at,
               e.id AS explanation_id,
               e.version,
               COALESCE(e.created_at, '') AS explanation_created_at,
               COALESCE(e.provider, '') AS provider,
               COALESCE(e.model, '') AS model,
               COALESCE(e.prompt_profile, '') AS prompt_profile,
               COALESCE(e.raw_text, '') AS raw_text,
               COALESCE({selected_fields['key_grammar']}, '') AS key_grammar,
               COALESCE({selected_fields['manual_original_text']}, '')
                   AS manual_original_text,
               COALESCE({selected_fields['manually_edited_at']}, '')
                   AS manually_edited_at,
               c.version_count
        FROM explanation_groups g
        JOIN selected e ON e.group_id = g.id AND e.selection_order = 1
        JOIN counts c ON c.group_id = g.id
        ORDER BY g.updated_at DESC, g.id DESC
        """
    ).fetchall()
    return [dict(row) for row in rows]


def details_by_group(connection: sqlite3.Connection) -> dict[int, dict[str, str]]:
    details: dict[int, dict[str, str]] = {}
    columns = table_columns(connection, "explanation_group_details")
    if columns:
        wanted = [
            name
            for name in ("chapter", "speaker", "tags", "added_to_anki_at")
            if name in columns
        ]
        select = ", ".join(f'COALESCE("{name}", \'\') AS "{name}"' for name in wanted)
        if select:
            for row in connection.execute(
                f"SELECT group_id, {select} FROM explanation_group_details"
            ):
                details[int(row["group_id"])] = {
                    name: str(row[name] or "") for name in wanted
                }

    if table_exists(connection, "explanation_group_tags") and table_exists(
        connection, "study_tags"
    ):
        for row in connection.execute(
            """
            SELECT gt.group_id, GROUP_CONCAT(t.name, ', ') AS tag_names
            FROM explanation_group_tags gt
            JOIN study_tags t ON t.id = gt.tag_id
            GROUP BY gt.group_id
            """
        ):
            item = details.setdefault(int(row["group_id"]), {})
            if not item.get("tags"):
                item["tags"] = str(row["tag_names"] or "")
    return details


def anki_by_group(connection: sqlite3.Connection) -> dict[int, dict[str, str]]:
    if not table_exists(connection, "explanation_group_anki_links"):
        return {}
    columns = table_columns(connection, "explanation_group_anki_links")
    wanted = [name for name in ("status", "note_id", "checked_at") if name in columns]
    select = ", ".join(f'COALESCE("{name}", \'\') AS "{name}"' for name in wanted)
    if not select:
        return {}
    return {
        int(row["group_id"]): {name: str(row[name] or "") for name in wanted}
        for row in connection.execute(
            f"SELECT group_id, {select} FROM explanation_group_anki_links"
        )
    }


def sections_by_explanation(
    connection: sqlite3.Connection,
) -> dict[int, list[dict[str, str]]]:
    if not table_exists(connection, "explanation_sections"):
        return {}
    sections: dict[int, list[dict[str, str]]] = {}
    for row in connection.execute(
        """
        SELECT explanation_id, COALESCE(section_key, '') AS section_key,
               COALESCE(heading, '') AS heading,
               COALESCE(content, '') AS content
        FROM explanation_sections
        ORDER BY explanation_id, sort_order, id
        """
    ):
        sections.setdefault(int(row["explanation_id"]), []).append(
            {
                "key": str(row["section_key"] or ""),
                "heading": str(row["heading"] or ""),
                "content": str(row["content"] or ""),
            }
        )
    return sections


def cached_recommendations(path: Path) -> dict[str, dict[str, Any]]:
    if not path.is_file():
        return {}
    connection = connect_read_only(path)
    try:
        if not table_exists(connection, "anki_candidate_recommendations"):
            return {}
        columns = table_columns(connection, "anki_candidate_recommendations")
        generated = (
            "COALESCE(generated_at, '')" if "generated_at" in columns else "''"
        )
        rows = connection.execute(
            f"""
            SELECT candidate_hash, recommended, score, reason, provider, model,
                   {generated} AS generated_at
            FROM anki_candidate_recommendations
            """
        ).fetchall()
        return {
            str(row["candidate_hash"]): {
                "recommended": bool(row["recommended"]),
                "score": int(row["score"] or 0),
                "reason": str(row["reason"] or ""),
                "provider": str(row["provider"] or ""),
                "model": str(row["model"] or ""),
                "generated_at": str(row["generated_at"] or ""),
            }
            for row in rows
        }
    finally:
        connection.close()


def anki_status(detail: dict[str, str], link: dict[str, str]) -> str:
    status = str(link.get("status", "")).strip().lower()
    if status in {"found", "linked", "exact_match", "added"}:
        return "Found in Anki"
    if detail.get("added_to_anki_at"):
        return "Added to Anki"
    if status and status not in {"not_checked", "missing", "not_found"}:
        return status.replace("_", " ").title()
    return "Not added"


def recommendation_label(item: dict[str, Any] | None) -> str:
    if item is None:
        return "Unassessed"
    return "Recommended" if item["recommended"] else "Not recommended"


def excel_value(value: Any) -> Any:
    """Keep arbitrary model output valid within Excel's cell constraints."""
    if not isinstance(value, str):
        return value
    value = ILLEGAL_XML_CHARACTERS.sub("", value)
    if len(value) > EXCEL_TEXT_LIMIT:
        marker = "\n\n[Truncated at the Excel cell limit]"
        value = value[: EXCEL_TEXT_LIMIT - len(marker)] + marker
    return value


def excel_datetime(value: Any) -> Any:
    """Return a native, timezone-free Excel datetime when possible.

    SQLite stores Study Library timestamps as ISO 8601 text. Excel has no
    timezone-aware datetime type, so preserve the timestamp's displayed local
    wall-clock value and remove only its offset metadata. Unrecognised values
    remain text rather than being lost.
    """
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, str) and value.strip():
        try:
            parsed = datetime.fromisoformat(value.strip())
        except ValueError:
            return excel_value(value)
    else:
        return excel_value(value)
    return parsed.replace(tzinfo=None)


def spreadsheet_value(header: str, value: Any) -> Any:
    if header in DATETIME_HEADERS:
        return excel_datetime(value)
    return excel_value(value)


def build_export_rows(
    connection: sqlite3.Connection,
    recommendations: dict[str, dict[str, Any]],
    recommendation_profile: str,
) -> tuple[list[list[Any]], list[list[Any]], list[list[Any]]]:
    groups = selected_explanations(connection)
    details = details_by_group(connection)
    links = anki_by_group(connection)
    section_map = sections_by_explanation(connection)

    explanation_rows: list[list[Any]] = []
    sentence_rows: list[list[Any]] = []
    vocabulary: OrderedDict[tuple[str, str], dict[str, Any]] = OrderedDict()

    for row in groups:
        group_id = int(row["group_id"])
        explanation_id = int(row["explanation_id"])
        profile = str(row["game_profile"] or "Unsorted") or "Unsorted"
        source = str(row["manual_original_text"] or row["source_japanese"] or "")
        group_detail = details.get(group_id, {})
        group_link = links.get(group_id, {})
        sections = section_map.get(explanation_id, [])
        headings = "\n".join(item["heading"] for item in sections)
        context_candidates = [
            item["content"]
            for item in sections
            if item["key"] in {"translation", "analysis", "vocabulary"}
        ]
        context = recommendation_language_context(
            next((value for value in context_candidates if value.strip()), "")
        )
        reason_language = recommendation_reason_language(
            str(row["prompt_profile"] or ""), headings, context
        )
        reason_context = context if reason_language.startswith("Match the language") else ""
        language_profile = (
            recommendation_profile + "\0reason-language=" + reason_language
        )
        if reason_context:
            language_profile += "\0" + reason_context
        sentence_key = candidate_hash(
            "sentence", source, recommendation_profile=language_profile
        )
        sentence_ai = recommendations.get(sentence_key)
        current_status = anki_status(group_detail, group_link)
        generated = str(row["group_updated_at"] or row["explanation_created_at"] or "")

        explanation_rows.append(
            [
                generated,
                profile,
                group_detail.get("chapter", ""),
                group_detail.get("speaker", ""),
                group_detail.get("tags", ""),
                source,
                str(row["key_grammar"] or ""),
                int(row["version_count"] or 0),
                current_status,
                group_detail.get("added_to_anki_at", ""),
                str(row["provider"] or ""),
                str(row["model"] or ""),
                str(row["prompt_profile"] or ""),
                f"v{int(row['version']):02d}",
                str(row["raw_text"] or ""),
                recommendation_label(sentence_ai),
                "" if sentence_ai is None else sentence_ai["score"],
                "" if sentence_ai is None else sentence_ai["reason"],
                "" if sentence_ai is None else sentence_ai["provider"],
                "" if sentence_ai is None else sentence_ai["model"],
                "" if sentence_ai is None else sentence_ai["generated_at"],
                group_id,
            ]
        )
        sentence_rows.append(
            [
                generated,
                profile,
                source,
                str(row["key_grammar"] or ""),
                int(row["version_count"] or 0),
                current_status,
                recommendation_label(sentence_ai),
                "" if sentence_ai is None else sentence_ai["score"],
                "" if sentence_ai is None else sentence_ai["reason"],
                "" if sentence_ai is None else sentence_ai["provider"],
                "" if sentence_ai is None else sentence_ai["model"],
                "" if sentence_ai is None else sentence_ai["generated_at"],
                group_id,
            ]
        )

        vocabulary_section = next(
            (item for item in sections if item["key"] == "vocabulary"), None
        )
        if vocabulary_section is None:
            continue
        seen_in_group: set[str] = set()
        for entry in parse_vocabulary(vocabulary_section["content"]):
            normalized = normalize_japanese(
                entry.front, remove_attached_readings=True
            )
            if not normalized or normalized in seen_in_group:
                continue
            seen_in_group.add(normalized)
            dedupe_key = (profile.casefold(), normalized)
            item = vocabulary.get(dedupe_key)
            if item is None:
                vocabulary[dedupe_key] = {
                    "display": entry.display,
                    "front": entry.front,
                    "back": entry.back,
                    "meaning": entry.meaning,
                    "occurrences": 1,
                    "profile": profile,
                    "latest": generated,
                    "anki_status": current_status,
                    "group_ids": [group_id],
                    "language_profile": language_profile,
                }
            else:
                item["occurrences"] += 1
                item["group_ids"].append(group_id)
                if generated > item["latest"]:
                    item.update(
                        display=entry.display,
                        front=entry.front,
                        back=entry.back,
                        meaning=entry.meaning,
                        latest=generated,
                        anki_status=current_status,
                        language_profile=language_profile,
                    )

    vocabulary_rows: list[list[Any]] = []
    for item in sorted(
        vocabulary.values(),
        key=lambda value: (str(value["latest"]), str(value["front"])),
        reverse=True,
    ):
        recommendation_key = candidate_hash(
            "vocabulary",
            str(item["front"]),
            str(item["back"]),
            str(item["language_profile"]),
        )
        ai = recommendations.get(recommendation_key)
        vocabulary_rows.append(
            [
                item["display"],
                item["front"],
                item["meaning"],
                item["back"],
                item["occurrences"],
                item["profile"],
                item["latest"],
                item["anki_status"],
                recommendation_label(ai),
                "" if ai is None else ai["score"],
                "" if ai is None else ai["reason"],
                "" if ai is None else ai["provider"],
                "" if ai is None else ai["model"],
                "" if ai is None else ai["generated_at"],
                ", ".join(str(value) for value in item["group_ids"]),
            ]
        )
    return explanation_rows, sentence_rows, vocabulary_rows


def add_sheet(
    workbook: Workbook,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    recommended_column: int | None = None,
) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 24

    for row_index, values in enumerate(rows, start=2):
        sheet.append([
            spreadsheet_value(headers[column_index], value)
            for column_index, value in enumerate(values)
        ])
        fill = ALT_FILL if row_index % 2 == 0 else None
        if recommended_column is not None and values[recommended_column - 1] == "Recommended":
            fill = RECOMMENDED_FILL
        for column_index, cell in enumerate(sheet[row_index]):
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if (
                headers[column_index] in DATETIME_HEADERS
                and isinstance(cell.value, datetime)
            ):
                cell.number_format = EXCEL_DATETIME_FORMAT
            if fill:
                cell.fill = fill

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    for column_index, header in enumerate(headers, start=1):
        values = [str(header)] + [
            str(row[column_index - 1] or "") for row in rows
        ]
        longest = max((max((len(line) for line in value.splitlines()), default=0) for value in values), default=8)
        width = min(MAX_COLUMN_WIDTH, max(10, longest + 2))
        sheet.column_dimensions[get_column_letter(column_index)].width = width


def export_workbook(arguments: argparse.Namespace) -> tuple[int, int, int]:
    database = Path(arguments.database)
    output = Path(arguments.output)
    preferences = Path(arguments.preferences_database)
    if not database.is_file():
        raise FileNotFoundError(f"Study Library database not found: {database}")

    recommendation_profile = recommendation_profile_signature(
        arguments.provider,
        arguments.model,
        arguments.learner_level,
        arguments.selection_style,
        arguments.focus_areas,
        arguments.additional_criteria,
    )
    recommendations = cached_recommendations(preferences)
    connection = connect_read_only(database)
    try:
        explanations, sentences, vocabulary = build_export_rows(
            connection, recommendations, recommendation_profile
        )
    finally:
        connection.close()

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = "JRPG Translator Study Library export"
    workbook.properties.creator = "JRPG Translator"
    workbook.properties.description = (
        "Read-only Study Library export with cached Anki recommendation data."
    )
    add_sheet(
        workbook,
        "Explanations",
        [
            "Date generated", "Profile", "Chapter", "Speaker", "Tags",
            "Japanese source", "Key grammar", "Versions", "Anki status",
            "Added to Anki at", "Provider", "Model", "Prompt",
            "Selected version", "Full explanation", "AI recommendation",
            "AI score", "AI reason", "AI provider", "AI model",
            "AI generated at", "Study Library ID",
        ],
        explanations,
        16,
    )
    add_sheet(
        workbook,
        "Sentence candidates",
        [
            "Date generated", "Profile", "Japanese source", "Key grammar",
            "Versions", "Anki status", "AI recommendation", "AI score",
            "AI reason", "AI provider", "AI model", "AI generated at",
            "Study Library ID",
        ],
        sentences,
        7,
    )
    add_sheet(
        workbook,
        "Vocabulary candidates",
        [
            "Vocabulary", "Japanese front", "Meaning / context", "Full entry",
            "Occurrences", "Profile", "Latest", "Anki status",
            "AI recommendation", "AI score", "AI reason", "AI provider",
            "AI model", "AI generated at", "Source Study Library IDs",
        ],
        vocabulary,
        9,
    )
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return len(explanations), len(sentences), len(vocabulary)


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--database", required=True)
    parser.add_argument("--preferences-database", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--provider", default="")
    parser.add_argument("--model", default="")
    parser.add_argument("--learner-level", default="intermediate")
    parser.add_argument("--selection-style", default="balanced")
    parser.add_argument("--focus-areas", default="")
    parser.add_argument("--additional-criteria-hex", default="")
    arguments = parser.parse_args()
    arguments.additional_criteria = decode_optional_hex(
        arguments.additional_criteria_hex
    )
    return arguments


def main() -> int:
    arguments = parse_arguments()
    counts = export_workbook(arguments)
    print(
        f"Exported {counts[0]} explanations, {counts[1]} sentence candidates, "
        f"and {counts[2]} vocabulary candidates to {arguments.output}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
